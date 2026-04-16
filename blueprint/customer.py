from flask import Blueprint,render_template,jsonify,redirect,url_for,session,request,render_template_string,flash
from datetime import datetime
import string
from exts import db
from models import SOHeaderModel,SOLineModel,MaterialModel,InventoryModel
import openpyxl
from snowflake import SnowflakeGenerator
from decorators import login_required,admin_required

customer_bp=Blueprint('customer',__name__,url_prefix='/customer')

#查找EXCEL元素函数
def find_excel_value(ws, target_labels):
    if isinstance(target_labels, str):
        target_labels = [target_labels]
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).strip().replace(" ", "")
                for label in target_labels:
                    clean_label = label.replace(" ", "")

                    if clean_label in cell_str:
                        target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                        return target_cell.value
    return None

#客户订单管理界面
@customer_bp.route('/somanage')
@login_required
@admin_required
def somanage():
    soheaders = SOHeaderModel.query.all()
    return render_template('somanage.html',soheaders=soheaders)

#客户订单详情界面
@customer_bp.route('/soinfo/<string:orderid>')
@login_required
@admin_required
def soinfo(orderid):
    soheader = SOHeaderModel.query.filter_by(orderid=orderid).first()
    if not soheader:
        flash('订单不存在！', 'error')
        return redirect(url_for('customer.somanage'))

    return render_template('soinfo.html', soheader=soheader)

#导入客户订单
@customer_bp.route('/import_so', methods=['POST'])
@login_required
@admin_required
def import_so():
    file = request.files.get('file')
    if not file:
        flash('请选择文件', 'error')
        return redirect(url_for('customer.somanage'))

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # 1. 生成雪花 ID
        gen = SnowflakeGenerator(42)
        orderid = str(next(gen))

        organization = find_excel_value(ws, ["单位名称", "单位名称："])
        if not organization:
            organization = find_excel_value(ws, ["需方单位", "需方单位："])
        
        if not organization:
            raise Exception("无法在模板中找到【单位名称】或【需方单位】信息")


        clientname = find_excel_value(ws, ["需方经办人", "需方经办人："])
        clientphone = find_excel_value(ws, ["需方联系电话", "需方联系电话："])

        address = find_excel_value(ws, ["收货详细地址", "收货详细地址：", "详细地址"])
        remarks = find_excel_value(ws, ["备注", "备注："])
        responsible = find_excel_value(ws, ["供方经办人", "供方经办人："])
        
        orderdate_val = find_excel_value(ws, ["下单日期", "下单日期："])
        deliverydate_val = find_excel_value(ws, ["要求交货日期", "要求交货日期："])


        if not clientname: raise Exception("未找到【需方经办人】")
        if not clientphone: raise Exception("未找到【需方联系电话】")
        if not address: raise Exception("未找到【收货详细地址】")
        if not responsible: raise Exception("未找到【供方经办人】")
        if not orderdate_val: raise Exception("未找到【下单日期】")
        if not deliverydate_val: raise Exception("未找到【要求交货日期】")


        def parse_date(date_val):
            if isinstance(date_val, datetime):
                return date_val
            try:
                return datetime.strptime(str(date_val), '%Y-%m-%d')
            except:
                return datetime.now()
        
        orderdate = parse_date(orderdate_val)
        deliverydate = parse_date(deliverydate_val)
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        deliverydate_midnight = deliverydate.replace(hour=0, minute=0, second=0, microsecond=0)
        
        if deliverydate_midnight <= today:
            raise Exception("导入失败：要求交货日期必须在今天之后！")

        header = SOHeaderModel(
            orderid=orderid,
            organization=organization,
            clientname=clientname,
            clientphone=str(clientphone),
            address=address,
            remarks=remarks,
            responsible=responsible,
            creationtime=datetime.now(),
            creator=session.get('username'),
            orderdate=orderdate,
            deliverydate=deliverydate,
            orderstatus='新建',
            completion=None
        )
        db.session.add(header)

        # 3. 解析行信息 (GSOLine)
        # 假设从第 5 行开始是数据，列对应关系依照模板图片：
        # B列(2): 产品代码, C列(3): 产品名称, D列(4): 规格型号, F列(6): 数量, G列(7): 单价
        start_row = 5
        row_index = start_row
        line_count = 0

        while True:
            # 获取 B 列的产品代码
            code_cell_val = ws.cell(row=row_index, column=2).value
            
            # 停止条件：如果产品代码为空，或者 A 列出现了 "合计" 字样
            check_end_val = ws.cell(row=row_index, column=1).value
            if not code_cell_val:
                # 再次确认一下是否是结束行（防止中间有空行但没结束的情况，这里简化处理为空即结束）
                break
            if check_end_val and "合计" in str(check_end_val):
                break

            materialcode = str(code_cell_val).strip()
            materialname = str(ws.cell(row=row_index, column=3).value).strip()
            spec = str(ws.cell(row=row_index, column=4).value).strip()
            
            try:
                qty = int(ws.cell(row=row_index, column=6).value)
                price = float(ws.cell(row=row_index, column=7).value)
            except (ValueError, TypeError):
                raise Exception(f"第 {row_index} 行的数据格式错误（数量或单价必须为数字）")

            # 数据库校验
            material = MaterialModel.query.filter_by(materialcode=materialcode).first()
            if not material:
                raise Exception(f"导入失败：第 {row_index} 行的产品代码 '{materialcode}' 在系统中不存在。")
            
            if material.materialdesc != materialname:
                raise Exception(f"导入失败：第 {row_index} 行的产品名称与系统不符（系统：{material.materialdesc}，Excel：{materialname}）")
            
            if material.specification != spec:
                raise Exception(f"导入失败：第 {row_index} 行的规格型号与系统不符（系统：{material.specification}，Excel：{spec}）")

            # 创建 Line 对象
            line = SOLineModel(
                orderid=orderid,
                materialcode=materialcode,
                quantity=qty,
                unitprice=price
            )
            db.session.add(line)
            
            line_count += 1
            row_index += 1

        if line_count == 0:
            raise Exception("未在模板中读取到任何有效的订单行数据。")

        # 全部成功，提交事务
        db.session.commit()
        flash('客户订单导入成功！', 'success')

    except Exception as e:
        # 任何一步出错，回滚所有操作
        db.session.rollback()
        print(e)
        flash(f'添加数据失败: {str(e)}', 'error')

    return redirect(url_for('customer.somanage'))

#交付订单操作
@customer_bp.route('/deliver_so', methods=['POST'])
@login_required
@admin_required
def deliver_so():
    data = request.get_json()
    orderid = data.get('orderid')
    
    soheader = SOHeaderModel.query.filter_by(orderid=orderid).first()
    if not soheader:
        return jsonify({'status': 'error', 'msg': '订单不存在！'})
        
    if soheader.orderstatus == '已交付':
        return jsonify({'status': 'error', 'msg': '该订单已交付！'})

    shortages = []
    
    # 1. 检查各行物料的总库存是否满足需求
    for line in soheader.solines:
        # 统计该物料在所有仓库中的库存总和
        total_inventory = db.session.query(db.func.sum(InventoryModel.quantity)).filter_by(materialcode=line.materialcode).scalar() or 0
        
        if total_inventory < line.quantity:
            shortage = line.quantity - total_inventory
            shortages.append(f"【{line.gmaterial.materialdesc}】(编码:{line.materialcode}) 需求:{line.quantity}，缺少:{shortage}")
            
    # 如果存在缺少的物料，则直接返回错误信息
    if shortages:
        return jsonify({'status': 'error', 'msg': '库存不足无法交付，具体缺少如下：\n' + '\n'.join(shortages)})
        
    # 2. 如果满足条件，扣减对应物料的库存（按现有库存从多到少优先扣除）
    try:
        for line in soheader.solines:
            remaining_to_deduct = line.quantity
            inventories = InventoryModel.query.filter_by(materialcode=line.materialcode).order_by(InventoryModel.quantity.desc()).all()
            
            for inv in inventories:
                if remaining_to_deduct <= 0:
                    break
                if inv.quantity >= remaining_to_deduct:
                    inv.quantity -= remaining_to_deduct
                    remaining_to_deduct = 0
                else:
                    remaining_to_deduct -= inv.quantity
                    inv.quantity = 0
                inv.altertime = datetime.now()
                
        # 更新订单状态与完成时间
        soheader.orderstatus = '已交付'
        soheader.completion = datetime.now()
        
        db.session.commit()
        return jsonify({'status': 'success', 'msg': '交付成功！'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'msg': f'数据库更新失败：{str(e)}'})